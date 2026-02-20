from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from datetime import datetime
from attendance.models import AttendanceSession, AttendanceRecord

def export_attendance_excel(classroom_id, start_date, end_date):
    """Export attendance data to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Header
    ws['A1'] = 'Attendance Report'
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f'Period: {start_date} to {end_date}'
    
    # Column headers
    headers = ['Date', 'Student Name', 'Subject', 'Status', 'Teacher']
    ws.append([])
    ws.append(headers)
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    for cell in ws[4]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    sessions = AttendanceSession.objects.filter(
        classroom_id=classroom_id,
        attendance_date__range=[start_date, end_date]
    ).select_related('teacher', 'subject')
    
    for session in sessions:
        records = AttendanceRecord.objects.filter(session=session).select_related('student')
        for record in records:
            ws.append([
                session.attendance_date.strftime('%Y-%m-%d'),
                record.student.full_name,
                session.subject.name,
                record.status.upper(),
                session.teacher.get_full_name()
            ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=attendance_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response

def export_attendance_pdf(classroom_id, start_date, end_date):
    """Export attendance data to PDF"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=attendance_report_{datetime.now().strftime("%Y%m%d")}.pdf'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(f'<b>Attendance Report</b><br/>{start_date} to {end_date}', styles['Title'])
    elements.append(title)
    
    # Data
    data = [['Date', 'Student', 'Subject', 'Status']]
    
    sessions = AttendanceSession.objects.filter(
        classroom_id=classroom_id,
        attendance_date__range=[start_date, end_date]
    ).select_related('subject')
    
    for session in sessions:
        records = AttendanceRecord.objects.filter(session=session).select_related('student')
        for record in records:
            data.append([
                session.attendance_date.strftime('%Y-%m-%d'),
                record.student.full_name,
                session.subject.name,
                record.status.upper()
            ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response
